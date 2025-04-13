'''
MCP Server for Excel Operations

This server provides tools to create, edit and manage Excel workbooks.
It's implemented using the Model Context Protocol (MCP) Python SDK.
'''

import os
import sys
import io
from mcp.server.fastmcp import FastMCP
from typing import Optional, List, Dict, Any, Union, Tuple

# 标记库是否已安装
openpyxl_installed = True

# 尝试导入openpyxl库，如果没有安装则标记为未安装但不退出
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.styles.colors import Color
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("警告: 未检测到openpyxl库，Excel功能将不可用")
    print("请使用以下命令安装: pip install openpyxl")
    openpyxl_installed = False

# 尝试导入Pandas库，用于数据处理
pandas_installed = True
try:
    import pandas as pd
    import numpy as np
except ImportError:
    print("警告: 未检测到pandas库，高级数据处理功能将受限")
    print("请使用以下命令安装: pip install pandas numpy")
    pandas_installed = False

# 创建一个MCP服务器，保持名称与配置文件一致
mcp = FastMCP("office editor")

@mcp.tool()
def create_excel_workbook(filename: str) -> str:
    """
    创建一个新的Excel工作簿。
    
    Args:
        filename: 要创建的文件名 (不需要包含.xlsx扩展名)
    
    Returns:
        包含操作结果的消息
    """
    # 检查是否安装了必要的库
    if not openpyxl_installed:
        return "错误: 无法创建Excel工作簿，请先安装openpyxl库: pip install openpyxl"
    
    # 确保文件名有.xlsx扩展名
    if not filename.lower().endswith('.xlsx'):
        filename += '.xlsx'
    
    # 从环境变量获取输出路径，如果未设置则使用默认桌面路径
    output_path = os.environ.get('OFFICE_EDIT_PATH')
    if not output_path:
        output_path = os.path.join(os.path.expanduser('~'), '桌面')
    
    # 创建完整的文件路径
    file_path = os.path.join(output_path, filename)
    
    try:
        # 创建输出目录（如果不存在）
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # 创建新的Excel工作簿
        wb = Workbook()
        
        # 保存工作簿
        wb.save(file_path)
        
        return f"成功在 {output_path} 创建了Excel工作簿: {filename}"
    except Exception as e:
        return f"创建Excel工作簿时出错: {str(e)}"

# 剩余代码省略，保持原样...

if __name__ == "__main__":
    # 运行MCP服务器
    print("启动Excel MCP服务器...")
    mcp.run()